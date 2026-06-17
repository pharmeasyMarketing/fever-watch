"""Build the analytics sign-off workbook: Fever_Watch_Score_Workbook.xlsx.

Combines, in one .xlsx:
  Overview, Data_Dictionary, Scores_2025, Scores_2026, Weather_2025, Weather_2026,
  Trends, Labs_2025, Labs_2026, Config_Refs.

The two Scores tabs keep the FULL in-cell formula build-up by merging the source
worksheet XML at the zip level (lossless, fast) rather than re-streaming 173K+
formula rows through openpyxl. Everything else is built with openpyxl, then the two
big worksheets are spliced in.

Run from the project root:  python scripts/build_score_workbook.py
"""
from __future__ import annotations

import csv
import json
import os
import re
import shutil
import time
import zipfile

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "data", "analytics")
OUT = os.path.join(OUT_DIR, "Fever_Watch_Score_Workbook.xlsx")

SRC_2025 = os.path.join(ROOT, "data", "backfill", "sheet", "raw_data_2025.xlsx")
SRC_2026 = os.path.join(ROOT, "data", "backfill", "sheet", "raw_data_2026_backfill.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F6F54")  # porcelain green
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color="1F6F54")
SUB_FONT = Font(bold=True, size=12, color="1F6F54")
THIN = Side(style="thin", color="D0D7D3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LOG = []
def log(msg):
    print(msg, flush=True)
    LOG.append(msg)


def style_header_row(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    # Use a string coordinate, not ws.cell(...), so we don't materialize an empty row below the header
    # (ws.cell(row=row+1, ...) would create a phantom blank row 2 that shifts the data down).
    ws.freeze_panes = "A%d" % (row + 1)


def autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------- Overview ----
def build_overview(wb, row_counts, live_labs_info, sampling_notes):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    A = lambda r, c=1: ws.cell(row=r, column=c)

    lines = []
    def add(text="", style=None, indent=0):
        lines.append((("    " * indent) + text, style))

    add("Fever Watch - End-to-End Risk-Score Workbook", "title")
    add("Analytics sign-off pack. Daily 0-100 risk score per city x disease, from three signals.", "muted")
    add("")
    add("WHAT THIS SCORES", "sub")
    add("A daily 0-100 risk indicator (NOT a diagnosis, case count, or medical advice) per city and disease")
    add("for India's top monsoon fevers: dengue (flagship), malaria, chikungunya, typhoid. The score is")
    add("always decomposable into its three signals; a forecast-only read (no lab data) is capped below HIGH.")
    add("")
    add("THE THREE SIGNALS", "sub")
    add("1. WEATHER (leading - conditions ahead). Environmental sub-score from NOAA CPC rainfall +")
    add("   NASA POWER temperature/humidity, shaped per disease family:")
    add("   mosquito (dengue/malaria/chikungunya) = ROUND((0.45*tempfit + 0.35*rain14_unit + 0.20*humidity_unit)*100)", indent=1)
    add("   waterborne (typhoid)                  = ROUND((0.6*rain7_unit + 0.4*rain14_unit)*100)", indent=1)
    add("   tempfit = unimodal response peaking at 29C, 0 outside 14-38C. *_unit = saturating 0..1 ramps.", indent=1)
    add("2. SEARCH / TRENDS (coincident - public concern). Google Trends state interest, floored:")
    add("   trends_score = MAX(4, MIN(100, state_interest)).", indent=1)
    add("3. LAB POSITIVITY (lagging - ground truth). PharmEasy / ThyroCare labs:")
    add("   positivity_pct = positives / tests_booked * 100", indent=1)
    add("   signal = MIN(100, ROUND(positivity_pct / ref * 100)), blanked when tests_booked < 30 (the gate).", indent=1)
    add("   ref is PER DISEASE: dengue 25, malaria 4, chikungunya 15, typhoid 45 (else 35).", indent=1)
    add("")
    add("THE ENSEMBLE (clubbing logic)", "sub")
    add("With positivity present (confirmed mode):")
    add("   score = ROUND( MIN(100, (0.30*weather + 0.22*trends + 0.48*positivity) * M) )", indent=1)
    add("   M = 1.08 if the three signals agree (max-min spread < 22) else 0.96 (the agree/disagree multiplier).", indent=1)
    add("Without positivity (forecast-only mode):")
    add("   score = ROUND( MIN(69, 0.60*weather + 0.40*trends) )", indent=1)
    add("   Capped at 69, one point below the HIGH band floor (70): a conditions-only read can never show HIGH.", indent=1)
    add("Per-city OVERALL headline (max-dominant blend):")
    add("   ROUND( 0.8*top-disease-score + 0.2*mean-of-the-rest ), with the driver disease named.", indent=1)
    add("Bands:  HIGH >= 70   MODERATE >= 45   LOW-MODERATE >= 25   LOW < 25.", indent=1)
    add("")
    add("DATA PROVENANCE", "sub")
    add("Rainfall:            NOAA CPC Global Unified Gauge-Based Analysis (gauge-based, US public domain).")
    add("Temperature/humidity: NASA POWER (CC0 / US public domain).")
    add("Search interest:     Google Trends (SerpApi / pytrends), state-level.")
    add("Lab positivity:      PharmEasy / ThyroCare labs (private feed; shown only as a city-level aggregate trend).")
    add("")
    add("REAL vs MOCK - read before sign-off", "sub")
    add("Scores_2025  : REAL labs (2025 historic PharmEasy feed). Full end-to-end build-up. Gold standard.", "warn")
    add("Scores_2026  : positivity is MOCK (placeholder), NOT the live lab feed - 2026-06-01..14 backfill.", "warn")
    add("Labs_2026    : the REAL live 2026 lab feed (city-aggregated), pulled separately for this pack.", "warn")
    add("All weather, trends and 2025 labs values are real.")
    add("")
    add("TABS IN THIS WORKBOOK", "sub")
    tab_desc = [
        ("Overview", "This methodology + provenance + real/mock notes (you are here)."),
        ("Data_Dictionary", "Every column in the Scores tabs explained (ported from docs/sheets_logging.md)."),
        ("Scores_2025", f"Full 2025 score build-up WITH in-cell formulas (real labs). {row_counts['Scores_2025']:,} data rows."),
        ("Scores_2026", f"Same build-up for 2026-06-01..14 (MOCK positivity). {row_counts['Scores_2026']:,} data rows."),
        ("Weather_2025", f"Raw weather value dump + per-family sub-scores. {row_counts['Weather_2025']:,} rows."),
        ("Weather_2026", f"Raw weather value dump + per-family sub-scores. {row_counts['Weather_2026']:,} rows."),
        ("Trends", f"Google Trends state interest + floored trends_score. {row_counts['Trends']:,} rows."),
        ("Labs_2025", f"2025 historic lab feed: tests, positives, pct, per-disease ref, gated signal. {row_counts['Labs_2025']:,} rows."),
        ("Labs_2026", live_labs_info),
        ("Config_Refs", "Weights, per-disease positivity refs, the 30-test gate, band thresholds, family weather weights."),
    ]
    for name, desc in tab_desc:
        add(f"{name}  -  {desc}")
    add("")
    add("BUILD NOTES (sampling / omissions)", "sub")
    if sampling_notes:
        for n in sampling_notes:
            add(n)
    else:
        add("None. The Scores tabs are the FULL formula build-up (no rows sampled or dropped).")
    add("All raw-dump tabs (Weather/Trends/Labs) are values-only by design (they carry no derived formulas).")
    add("")
    add(f"Generated {time.strftime('%Y-%m-%d %H:%M')} by scripts/build_score_workbook.py", "muted")

    r = 1
    for text, style in lines:
        cell = A(r)
        cell.value = text
        if style == "title":
            cell.font = TITLE_FONT
        elif style == "sub":
            cell.font = SUB_FONT
        elif style == "warn":
            cell.font = Font(bold=True, color="B5520E")
        elif style == "muted":
            cell.font = Font(italic=True, color="6B7B73", size=10)
        else:
            cell.font = Font(size=11)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        r += 1
    ws.column_dimensions["A"].width = 120
    log(f"  Overview: {r-1} lines")


# --------------------------------------------------------- Data_Dictionary ----
DICT = [
    ['Column', 'What it is / how it is derived'],
    ['date', 'UTC date of the run (grid generated_at).'],
    ['run_id', 'Run id; groups one pipeline run.'],
    ['city', 'City name.'],
    ['state', 'State / UT.'],
    ['disease', 'Disease label, or OVERALL for the city headline row.'],
    ['family', 'Weather model family: mosquito / waterborne (selects how weather is shaped).'],
    ['temp_c', 'Trailing mean air temperature (C), NASA POWER. Input to weather_score.'],
    ['humidity_pct', 'Trailing mean relative humidity (%), NASA POWER. Input to weather_score.'],
    ['rain_7d_mm', 'Rainfall over the last 7 days (mm), NOAA CPC (gauge-based, US public domain). Input to weather_score.'],
    ['rain_14d_mm', 'Rainfall over the last 14 days (mm), NOAA CPC; the lagged breeding signal.'],
    ['weather_score', 'FORMULA (build-up): 0-100 breeding/transmission favourability, derived in-cell from temp_c/humidity_pct/rain_7d_mm/rain_14d_mm. mosquito = (0.45*tempfit + 0.35*rain14_unit + 0.20*humidity_unit)*100; waterborne = (0.6*rain7_unit + 0.4*rain14_unit)*100. tempfit = unimodal peak at 29C (0 outside 14-38C); *_unit = saturating ramps (config/scoring.json).'],
    ['trends_score', 'FORMULA: MAX(4, MIN(100, trends_state_interest)) - the state Google Trends index for trends_keywords, floored at 4 and capped at 100.'],
    ['trends_keywords', 'The search terms (OR-joined) whose combined interest is trends_score.'],
    ['news_spike', 'TRUE if national interest spiked recently (news-driven); trends is down-weighted in forecast mode when TRUE.'],
    ['positivity', 'FORMULA: 0-100 PharmEasy lab positivity (lagging ground truth) = MIN(100, ROUND(positivity_pct/ref*100)), gated to blank when tests_booked < 30 (forecast-only). ref is PER DISEASE (dengue 25, malaria 4, chikungunya 15, typhoid 45, else 35), looked up in-cell from the disease in column E.'],
    ['w_weather', 'Weight (%) on weather_score in the blend (30 confirmed / 60 forecast).'],
    ['w_trends', 'Weight (%) on trends_score (22 confirmed / 40 forecast).'],
    ['w_positivity', 'Weight (%) on positivity (48 confirmed / 0 forecast).'],
    ['confidence', 'FORMULA: Forecast only if no positivity; else High if the three signals agree (max-min < 22) else Moderate; downgraded one step (High->Moderate, Moderate->Low) when stale=TRUE.'],
    ['score', 'FORMULA. confirmed = (w_weather*weather + w_trends*trends + w_positivity*positivity) x1.08 if signals agree else x0.96, capped 100. forecast = weather+trends only, capped 69. OVERALL = 0.8*top disease + 0.2*mean of the rest.'],
    ['band', 'FORMULA: HIGH >=70, MODERATE >=45, LOW-MODERATE >=25, LOW <25 (from score).'],
    ['mode', 'FORMULA: confirmed (positivity present) or forecast (capped, no positivity); OVERALL rows = blend.'],
    ['trends_state_interest', 'Raw Google Trends interest (0-100) for the city state (or the disease national mean if the state has no row), BEFORE the floor. trends_score = MAX(4, MIN(100, this)).'],
    ['weather_fresh', 'How fresh the weather data is: fresh (today) / carried Nd / stale Nd. From weather.json generated_at.'],
    ['trends_fresh', 'How fresh this disease trends data is (carry-forward on a SerpApi failure): fresh / carried Nd / stale Nd / unknown.'],
    ['stale', 'TRUE if any signal is older than stale_days (config/consolidation.json) - the cell used a carried-forward reading and its confidence is downgraded one step.'],
    ['weather_source', 'Provenance of the weather inputs: rainfall from NOAA CPC; temperature and humidity from NASA POWER.'],
    ['tests_booked', 'Aggregate lab tests for this city/disease over the trailing window (config window_days), PharmEasy/ThyroCare feed. Raw input to positivity; logged here only (never in the public site). [2025 Scores tab only]'],
    ['positives', 'Aggregate positive results over the same window. positivity_pct = positives / tests_booked * 100. [2025 Scores tab only]'],
    ['positivity_pct', 'FORMULA: positives(AC)/tests_booked(AB)*100. Completes the build-up: tests_booked, positives -> positivity_pct -> positivity (O) -> score (T). [2025 Scores tab only]'],
]
def build_dictionary(wb):
    ws = wb.create_sheet("Data_Dictionary")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Data Dictionary - columns in Scores_2025 / Scores_2026").font = TITLE_FONT
    ws.append([])
    ws.append(DICT[0])
    hr = ws.max_row
    style_header_row(ws, 2, row=hr)
    for row in DICT[1:]:
        ws.append(row)
    for r in range(hr + 1, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
        for c in (1, 2):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=r, column=c).border = BORDER
    autowidth(ws, [22, 150])
    log(f"  Data_Dictionary: {len(DICT)-1} columns documented")


# ------------------------------------------------------------- Weather dump ---
def build_weather(wb, year, src_json):
    data = json.load(open(src_json, encoding="utf-8"))
    cities = json.load(open(os.path.join(ROOT, "config", "cities.json"), encoding="utf-8"))["cities"]
    name_by_id = {c["id"]: c["name"] for c in cities}
    state_by_id = {c["id"]: c["state"] for c in cities}
    ws = wb.create_sheet(f"Weather_{year}")
    headers = ["city_id", "city", "state", "date", "temp_c", "humidity_pct",
               "rain_7d_mm", "rain_14d_mm", "weather_mosquito", "weather_waterborne", "provider"]
    ws.append(headers)
    style_header_row(ws, len(headers))
    provider = data.get("provider", "cpc")
    by_date = data.get("by_date", {})
    n = 0
    for date in sorted(by_date.keys()):
        for cid in sorted(by_date[date].keys()):
            cell = by_date[date][cid]
            agg = cell.get("agg", {})
            fam = cell.get("families", {})
            ws.append([
                cid, name_by_id.get(cid, cid), state_by_id.get(cid, ""), date,
                agg.get("temp_mean_c"), agg.get("humidity_pct"),
                agg.get("rain_7d_mm"), agg.get("rain_14d_mm"),
                fam.get("mosquito"), fam.get("waterborne"), provider,
            ])
            n += 1
    autowidth(ws, [14, 18, 18, 12, 9, 12, 11, 12, 16, 17, 9])
    log(f"  Weather_{year}: {n:,} rows")
    return n


# -------------------------------------------------------------- Trends dump ---
def build_trends(wb):
    data = json.load(open(os.path.join(ROOT, "data", "backfill", "trends_history.json"), encoding="utf-8"))
    diseases = data.get("diseases", {})
    label = {"dengue": "Dengue", "malaria": "Malaria", "chikungunya": "Chikungunya", "typhoid": "Typhoid"}
    ws = wb.create_sheet("Trends")
    headers = ["disease", "state", "week_start", "state_interest", "trends_score"]
    ws.append(headers)
    style_header_row(ws, len(headers))
    n = 0
    for did in sorted(diseases.keys()):
        node = diseases[did]
        for state in sorted(node.get("by_state", {}).keys()):
            for wk in node["by_state"][state]:
                week_start, value = wk[0], wk[1]
                # trends_score = MAX(4, MIN(100, state_interest)) as an in-cell formula
                r = ws.max_row + 1
                ws.append([label.get(did, did), state, week_start, value, None])
                ws.cell(row=r, column=5).value = f"=MAX(4,MIN(100,D{r}))"
                n += 1
    autowidth(ws, [14, 28, 13, 15, 13])
    log(f"  Trends: {n:,} rows")
    return n


# ------------------------------------------------------------- Labs_2025 dump -
def build_labs_2025(wb):
    path = os.path.join(ROOT, "data", "lab_feed_2025_historic.csv")
    ws = wb.create_sheet("Labs_2025")
    n = 0
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        header = next(r)
        # add a ref column for clarity
        out_header = header + ["ref_positivity_pct"]
        ws.append(out_header)
        style_header_row(ws, len(out_header))
        ref_by = {"dengue": 25, "malaria": 4, "chikungunya": 15, "typhoid": 45}
        di = header.index("disease")
        for row in r:
            disease = row[di].strip().lower()
            ref = ref_by.get(disease, 35)
            # coerce numerics
            out = list(row)
            for idx in (header.index("tests_booked"), header.index("positives")):
                try:
                    out[idx] = int(out[idx])
                except (ValueError, TypeError):
                    pass
            for idx in (header.index("positivity_pct"), header.index("positivity_signal")):
                if out[idx] not in ("", None):
                    try:
                        out[idx] = float(out[idx])
                    except (ValueError, TypeError):
                        pass
            try:
                out[header.index("season_week")] = int(out[header.index("season_week")])
            except (ValueError, TypeError):
                pass
            ws.append(out + [ref])
            n += 1
    autowidth(ws, [12, 12, 16, 14, 13, 10, 14, 16, 17])
    log(f"  Labs_2025: {n:,} rows")
    return n


# ------------------------------------------------------------- Labs_2026 dump -
def build_labs_2026(wb, live_json):
    ws = wb.create_sheet("Labs_2026")
    if live_json and os.path.exists(live_json):
        blob = json.load(open(live_json, encoding="utf-8"))
        rows = blob["rows"]
        meta = blob.get("meta", {})
        note = (f"Live 2026 PharmEasy/ThyroCare lab feed via the Google Sheets API "
                f"(service account). City-aggregated over a trailing {meta.get('window_days')}-day window; "
                f"NO per-patient rows. {meta.get('raw_rows_pulled'):,} raw daily rows -> "
                f"{len(rows):,} (city x disease) pairs. signal is blank when tests_booked < 30 (the gate).")
        ws.cell(row=1, column=1, value=note).font = Font(italic=True, color="B5520E", size=10)
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=False)
        ws.append([])
        headers = ["city_id", "city", "state", "disease", "tests_booked", "positives",
                   "positivity_pct", "ref_positivity_pct", "window_days", "gated_below_30", "positivity_signal"]
        ws.append(headers)
        hr = ws.max_row
        style_header_row(ws, len(headers), row=hr)
        ld = {"dengue": "Dengue", "malaria": "Malaria", "chikungunya": "Chikungunya", "typhoid": "Typhoid"}
        for row in rows:
            ws.append([
                row["city_id"], row["city"], row["state"], ld.get(row["disease"], row["disease"]),
                row["tests_booked"], row["positives"], row["positivity_pct"], row["ref_pct"],
                row["window_days"], "yes" if row["gated"] else "no", row["signal"],
            ])
        autowidth(ws, [14, 18, 18, 13, 12, 10, 14, 17, 12, 14, 17])
        log(f"  Labs_2026: {len(rows):,} rows (LIVE feed included)")
        return len(rows), True
    else:
        ws.cell(row=1, column=1,
                value="Live 2026 lab feed requires the service-account key; not available in this build.").font = Font(bold=True, color="B5520E")
        ws.column_dimensions["A"].width = 100
        log("  Labs_2026: note only (live feed unavailable)")
        return 0, False


# ----------------------------------------------------------- Config_Refs tab --
def build_config_refs(wb):
    ws = wb.create_sheet("Config_Refs")
    ws.sheet_view.showGridLines = False
    r = 1
    def title(t):
        nonlocal r
        ws.cell(row=r, column=1, value=t).font = SUB_FONT
        r += 1
    def table(headers, rows, widths=None):
        nonlocal r
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=j, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(vertical="center", wrap_text=True)
        r += 1
        for row in rows:
            for j, v in enumerate(row, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.border = BORDER
                c.alignment = Alignment(vertical="center")
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="Config Reference - scoring constants").font = TITLE_FONT
    r += 2

    title("Ensemble weights")
    table(["Mode", "Weather", "Trends", "Positivity", "Notes"],
          [["Confirmed (positivity present)", 0.30, 0.22, 0.48, "score = ROUND(MIN(100, blend * agree_mult))"],
           ["Forecast-only (no positivity)", 0.60, 0.40, "n/a", "score = ROUND(MIN(69, blend)); capped below HIGH"]])

    title("Agreement multiplier (confirmed mode only)")
    table(["Condition", "Multiplier"],
          [["Signals agree: max-min spread < 22", 1.08],
           ["Signals disagree: spread >= 22", 0.96]])

    title("Forecast cap")
    table(["Constant", "Value"], [["score_cap (forecast-only)", 69], ["HIGH band floor", 70]])

    title("Per-disease positivity reference (% that maps to a full 100 signal)")
    table(["Disease", "ref_positivity_pct"],
          [["Dengue", 25], ["Malaria", 4], ["Chikungunya", 15], ["Typhoid", 45], ["Any other (fallback)", 35]])

    title("Lab confidence gate")
    table(["Constant", "Value", "Effect"],
          [["min_tests", 30, "positivity signal is blank (forecast-only) below this many tests"],
           ["window_days", 14, "trailing days of lab rows summed before the gate"]])

    title("Band thresholds (from final score)")
    table(["Band", "Min score", "Colour"],
          [["HIGH", 70, "#d64545"], ["MODERATE", 45, "#d98a2b"],
           ["LOW-MODERATE", 25, "#c2a93a"], ["LOW", 0, "#3f9d6f"]])

    title("Family weather weights + shaping")
    table(["Family", "Weights", "Shaping constants"],
          [["mosquito (dengue/malaria/chikungunya)", "temp 0.45, rain_lagged 0.35, humidity 0.20",
            "temp_optimal 29C, range 14-38C, width 13, rain14 saturation 60mm, humidity 40-90%"],
           ["waterborne (typhoid)", "rain_recent 0.60, rain_lagged 0.40",
            "rain7 + rain14, saturation 80mm; temperature minor"]])

    title("City headline blend")
    table(["Constant", "Value"],
          [["top_weight (driver disease)", 0.8], ["rest_weight (mean of the rest)", 0.2]])

    title("Data provenance + cadence")
    table(["Signal", "Source", "Cadence"],
          [["Rainfall", "NOAA CPC Global Unified Gauge-Based Analysis (US public domain)", "daily"],
           ["Temperature / humidity", "NASA POWER (CC0)", "daily"],
           ["Search interest", "Google Trends (SerpApi), state-level", "weekly timeseries / daily snapshot"],
           ["Lab positivity", "PharmEasy / ThyroCare labs (private feed)", "daily"]])

    autowidth(ws, [40, 38, 60, 14, 50])
    log("  Config_Refs: built")


# ------------------------------------------------ splice big Scores tabs in ---
def count_rows_xlsx(path):
    """Exact data-row count via the worksheet XML (fast: streams the part, tracks the
    last <row r="N"> ref). Avoids the multi-minute openpyxl load of the 36MB workbook."""
    with zipfile.ZipFile(path) as z:
        part = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet")][0]
        last = 0
        rx = re.compile(rb'<row[^>]*\br="(\d+)"')
        with z.open(part) as fh:
            tail = b""
            while True:
                chunk = fh.read(1 << 20)  # 1 MB
                if not chunk:
                    break
                buf = tail + chunk
                for m in rx.finditer(buf):
                    v = int(m.group(1))
                    if v > last:
                        last = v
                tail = buf[-64:]  # carry a small overlap so a split tag is not missed
    return max(0, last - 1)  # minus the header row


def splice_scores(base_path, src_path, sheet_title):
    """Add the source workbook's single worksheet into base_path as `sheet_title`,
    preserving every in-cell formula. Done at the zip/XML level (lossless, fast).

    openpyxl has already created an empty placeholder worksheet named `sheet_title`
    in base_path. We replace that worksheet's XML part with the source worksheet XML,
    and merge the source sharedStrings into the base sharedStrings (re-indexing).
    """
    import xml.etree.ElementTree as ET
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", NS)

    # 1. find which sheetN.xml the placeholder maps to in the base workbook
    with zipfile.ZipFile(base_path) as bz:
        wb_xml = bz.read("xl/workbook.xml").decode("utf-8")
        rels_xml = bz.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        base_ss = bz.read("xl/sharedStrings.xml").decode("utf-8") if "xl/sharedStrings.xml" in bz.namelist() else None
        base_names = bz.namelist()

    # sheet name -> r:id (attribute order is not guaranteed, so match each attr independently)
    rid = None
    for sm in re.finditer(r"<sheet\b[^>]*/>", wb_xml):
        tag = sm.group(0)
        nm = re.search(r'\bname="([^"]+)"', tag)
        if nm and nm.group(1) == sheet_title:
            ridm = re.search(r'\br:id="([^"]+)"', tag)
            rid = ridm.group(1)
            break
    if rid is None:
        raise RuntimeError(f"could not find sheet '{sheet_title}' in workbook.xml")
    # r:id -> Target (Id and Target may appear in either order)
    target = None
    for rm in re.finditer(r"<Relationship\b[^>]*/>", rels_xml):
        tag = rm.group(0)
        idm = re.search(r'\bId="([^"]+)"', tag)
        if idm and idm.group(1) == rid:
            tm = re.search(r'\bTarget="([^"]+)"', tag)
            target = tm.group(1)  # e.g. /xl/worksheets/sheet3.xml or worksheets/sheet3.xml
            break
    if target is None:
        raise RuntimeError(f"could not resolve relationship {rid} in workbook.xml.rels")
    # Targets may be absolute (/xl/worksheets/sheetN.xml) or relative (worksheets/sheetN.xml)
    if target.startswith("/"):
        target_part = target.lstrip("/")            # /xl/worksheets/sheet3.xml -> xl/worksheets/sheet3.xml
    else:
        target_part = "xl/" + target                 # worksheets/sheet3.xml     -> xl/worksheets/sheet3.xml

    # 2. read the source worksheet + its shared strings
    with zipfile.ZipFile(src_path) as sz:
        src_sheet_part = [n for n in sz.namelist() if n.startswith("xl/worksheets/sheet")][0]
        src_sheet_xml = sz.read(src_sheet_part).decode("utf-8")
        src_ss_xml = sz.read("xl/sharedStrings.xml").decode("utf-8") if "xl/sharedStrings.xml" in sz.namelist() else None

    # 3. parse base + source shared strings; build the merged table + remap
    def parse_sst(xml):
        if not xml:
            return []
        root = ET.fromstring(xml)
        return list(root)  # list of <si> elements

    base_si = parse_sst(base_ss)
    src_si = parse_sst(src_ss_xml)

    # index base <si> by their serialized form so we can dedupe
    def si_key(si):
        return ET.tostring(si, encoding="unicode")
    base_keys = {si_key(si): i for i, si in enumerate(base_si)}
    merged = list(base_si)
    remap = {}  # src index -> merged index
    for i, si in enumerate(src_si):
        k = si_key(si)
        if k in base_keys:
            remap[i] = base_keys[k]
        else:
            base_keys[k] = len(merged)
            remap[i] = len(merged)
            merged.append(si)

    # 4. rewrite the source sheet's string-cell references (t="s") through remap
    #    cells look like <c r="C2" t="s"><v>123</v></c>
    def remap_cell(mobj):
        head, body = mobj.group(1), mobj.group(2)
        def repl_v(vm):
            return "<v>%d</v>" % remap.get(int(vm.group(1)), int(vm.group(1)))
        body2 = re.sub(r"<v>(\d+)</v>", repl_v, body)
        return head + body2 + "</c>"
    new_sheet_xml = re.sub(r'(<c\b[^>]*\bt="s"[^>]*>)(.*?)</c>', remap_cell, src_sheet_xml, flags=re.DOTALL)

    # 5. serialize the merged shared strings table
    sst_root = ET.Element("{%s}sst" % NS)
    sst_root.set("count", str(len(merged)))
    sst_root.set("uniqueCount", str(len(merged)))
    for si in merged:
        sst_root.append(si)
    merged_ss_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + ET.tostring(sst_root, encoding="unicode")

    # 6. rewrite the zip: replace the placeholder worksheet part + sharedStrings part
    tmp = base_path + ".tmp"
    with zipfile.ZipFile(base_path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
        for item in zin.infolist():
            name = item.filename
            if name == target_part:
                zout.writestr(item, new_sheet_xml)
            elif name == "xl/sharedStrings.xml":
                zout.writestr(item, merged_ss_xml)
            else:
                zout.writestr(item, zin.read(name))
        if "xl/sharedStrings.xml" not in zin.namelist():
            zout.writestr("xl/sharedStrings.xml", merged_ss_xml)
    shutil.move(tmp, base_path)
    log(f"  spliced {sheet_title}: worksheet XML + {len(merged):,} shared strings")


def main():
    t0 = time.time()
    os.makedirs(OUT_DIR, exist_ok=True)
    log("Building Fever Watch score workbook...")

    # exact Scores row counts up front (for the Overview tab)
    log("Counting Scores rows (this reads the big xlsx headers)...")
    n2025 = count_rows_xlsx(SRC_2025)
    n2026 = count_rows_xlsx(SRC_2026)
    log(f"  Scores_2025: {n2025:,} rows | Scores_2026: {n2026:,} rows")

    live_json = os.path.join(OUT_DIR, "_labs_2026_live.json")
    live_present = os.path.exists(live_json)
    n_live = 0
    if live_present:
        n_live = len(json.load(open(live_json, encoding="utf-8"))["rows"])

    # ---- build the small tabs with openpyxl ----
    wb = openpyxl.Workbook()
    row_counts = {"Scores_2025": n2025, "Scores_2026": n2026}

    # placeholders for the two big sheets (filled by splice later), positioned in order
    # build order matters for final tab order: Overview, Data_Dictionary, [Scores_2025],
    # [Scores_2026], Weather_2025, Weather_2026, Trends, Labs_2025, Labs_2026, Config_Refs
    overview = wb.active  # becomes "Overview"
    build_dictionary(wb)
    sc25 = wb.create_sheet("Scores_2025")  # placeholder
    sc25.cell(row=1, column=1, value="(placeholder - replaced by full formula build-up at splice time)")
    sc26 = wb.create_sheet("Scores_2026")  # placeholder
    sc26.cell(row=1, column=1, value="(placeholder - replaced by full formula build-up at splice time)")
    row_counts["Weather_2025"] = build_weather(wb, 2025, os.path.join(ROOT, "data", "backfill", "weather_2025.json"))
    row_counts["Weather_2026"] = build_weather(wb, 2026, os.path.join(ROOT, "data", "backfill", "weather_2026.json"))
    row_counts["Trends"] = build_trends(wb)
    row_counts["Labs_2025"] = build_labs_2025(wb)
    n_labs26, live_included = build_labs_2026(wb, live_json if live_present else None)
    row_counts["Labs_2026"] = n_labs26
    build_config_refs(wb)

    live_info = (f"Live 2026 lab feed (city-aggregated). {n_labs26:,} (city x disease) pairs."
                 if live_included else
                 "Live 2026 lab feed requires the service-account key; not available in this build.")
    sampling_notes = []  # full formula copy, no sampling
    build_overview(wb, row_counts, live_info, sampling_notes)

    wb.save(OUT)
    log(f"Base workbook saved ({os.path.getsize(OUT)/1e6:.1f} MB). Splicing Scores tabs...")

    # ---- splice the two big formula worksheets in ----
    splice_scores(OUT, SRC_2025, "Scores_2025")
    splice_scores(OUT, SRC_2026, "Scores_2026")

    size = os.path.getsize(OUT)
    log(f"DONE in {time.time()-t0:.0f}s. {OUT} = {size/1e6:.1f} MB")
    return row_counts, n_live, live_included, size


if __name__ == "__main__":
    main()
